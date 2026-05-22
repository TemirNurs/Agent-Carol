#!/usr/bin/env python3
"""
CCF Estimate Excel Exporter
Exports estimate JSON to Excel matching CCF template format.
Usage: python estimate_to_xlsx.py --estimate estimate.json --output estimate.xlsx
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Arial", bold=True, size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=14)
SUBTITLE_FONT = Font(name="Arial", bold=True, size=10, color="333333")
DATA_FONT = Font(name="Arial", size=10)
CURRENCY_FMT = '$#,##0.00'
PCT_FMT = '0.0%'
NUM_FMT = '#,##0.00'

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT_WHITE = Font(name="Arial", bold=True, size=10, color="FFFFFF")
TOTAL_FILL = PatternFill("solid", fgColor="D6E4F0")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="999999"),
)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def style_total_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = TOTAL_FILL


def write_summary_sheet(wb, estimate):
    ws = wb.active
    ws.title = "Estimate Summary"

    proj = estimate.get("project", {})
    totals = estimate.get("totals", {})
    metrics = estimate.get("metrics", {})
    categories = estimate.get("category_subtotals", {})

    # Title
    ws["A1"] = "PAINTING ESTIMATE"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = proj.get("name", "")
    ws["A2"].font = SUBTITLE_FONT
    ws["A4"] = f"Bid Date: {proj.get('bid_date', '')} | Prepared by: Carolina Commercial Finishes"
    ws["A4"].font = DATA_FONT

    # Summary table
    row = 7
    headers = ["SCOPE CATEGORY", "LABOR HRS", "LABOR $", "MATERIALS $", "DIRECT COST", "% OF TOTAL"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=row, column=col, value=h)
    style_header_row(ws, row, len(headers))

    row = 8
    direct_cost = totals.get("direct_cost", 0)
    for cat_name, cat_data in categories.items():
        ws.cell(row=row, column=1, value=cat_name).font = DATA_FONT
        ws.cell(row=row, column=2, value=cat_data["labor_hours"]).number_format = NUM_FMT
        ws.cell(row=row, column=3, value=cat_data["labor_cost"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=4, value=cat_data["material_cost"]).number_format = CURRENCY_FMT
        ws.cell(row=row, column=5, value=cat_data["subtotal"]).number_format = CURRENCY_FMT
        pct = cat_data["subtotal"] / direct_cost if direct_cost > 0 else 0
        ws.cell(row=row, column=6, value=pct).number_format = PCT_FMT
        row += 1

    # Totals
    ws.cell(row=row, column=1, value="DIRECT COST TOTALS").font = HEADER_FONT
    ws.cell(row=row, column=2, value=totals["labor_hours"]).number_format = NUM_FMT
    ws.cell(row=row, column=3, value=totals["labor_cost"]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=4, value=totals["material_cost"]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=5, value=totals["direct_cost"]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=6, value=1.0).number_format = PCT_FMT
    style_total_row(ws, row, 6)

    row += 2
    ws.cell(row=row, column=1, value="Direct Cost").font = DATA_FONT
    ws.cell(row=row, column=5, value=totals["direct_cost"]).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value=f"Overhead ({proj.get('overhead_pct', 0.12):.0%})").font = DATA_FONT
    ws.cell(row=row, column=5, value=totals["overhead"]).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value=f"Markup ({proj.get('markup_pct', 0.20):.0%})").font = DATA_FONT
    ws.cell(row=row, column=5, value=totals["markup"]).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="BID PRICE").font = Font(name="Arial", bold=True, size=12)
    ws.cell(row=row, column=5, value=totals["bid_price"]).number_format = CURRENCY_FMT
    ws.cell(row=row, column=5).font = Font(name="Arial", bold=True, size=12)
    style_total_row(ws, row, 6)

    row += 2
    ws.cell(row=row, column=1, value="KEY METRICS").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Total Paintable SF:").font = DATA_FONT
    ws.cell(row=row, column=2, value=metrics.get("total_sf", 0)).number_format = '#,##0'
    row += 1
    ws.cell(row=row, column=1, value="Blended Rate ($/SF):").font = DATA_FONT
    ws.cell(row=row, column=2, value=metrics.get("blended_rate_per_sf", 0)).number_format = CURRENCY_FMT
    row += 1
    ws.cell(row=row, column=1, value="Total Labor Hours:").font = DATA_FONT
    ws.cell(row=row, column=2, value=totals.get("labor_hours", 0)).number_format = NUM_FMT
    row += 1
    ws.cell(row=row, column=1, value="Est. Duration (2-man crew):").font = DATA_FONT
    ws.cell(row=row, column=2, value=f"{metrics.get('crew_days_2_man', 0)} days")
    row += 1
    ws.cell(row=row, column=1, value="Est. Duration (3-man crew):").font = DATA_FONT
    ws.cell(row=row, column=2, value=f"{metrics.get('crew_days_3_man', 0)} days")
    row += 1
    ws.cell(row=row, column=1, value="Labor Rate (burdened):").font = DATA_FONT
    ws.cell(row=row, column=2, value=f"${proj.get('labor_rate', 28):.2f}/hr")

    # Column widths
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 12


def write_detail_sheet(wb, sheet_name, line_items):
    """Write a detail sheet for a category of line items."""
    ws = wb.create_sheet(sheet_name)

    headers = ["Area", "Task", "Qty", "Unit", "Method", "Coats", "Prod Rate",
               "Labor Hrs", "Labor $", "Material $", "Total $", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))

    for i, li in enumerate(line_items, 2):
        ws.cell(row=i, column=1, value=li.get("area", "")).font = DATA_FONT
        ws.cell(row=i, column=2, value=li.get("task", "")).font = DATA_FONT
        ws.cell(row=i, column=3, value=li.get("quantity", 0)).number_format = '#,##0'
        ws.cell(row=i, column=4, value=li.get("unit", "")).font = DATA_FONT
        ws.cell(row=i, column=5, value=li.get("method", "")).font = DATA_FONT
        ws.cell(row=i, column=6, value=li.get("coats", 1))
        ws.cell(row=i, column=7, value=li.get("prod_rate", "")).number_format = '#,##0'
        ws.cell(row=i, column=8, value=li.get("labor_hours", 0)).number_format = NUM_FMT
        ws.cell(row=i, column=9, value=li.get("labor_cost", 0)).number_format = CURRENCY_FMT
        ws.cell(row=i, column=10, value=li.get("material_cost", 0)).number_format = CURRENCY_FMT
        ws.cell(row=i, column=11, value=li.get("subtotal", 0)).number_format = CURRENCY_FMT
        ws.cell(row=i, column=12, value=li.get("notes", "")).font = DATA_FONT

    # Totals row
    total_row = len(line_items) + 2
    ws.cell(row=total_row, column=1, value=f"TOTAL — {sheet_name}").font = HEADER_FONT
    ws.cell(row=total_row, column=8, value=sum(li.get("labor_hours", 0) for li in line_items)).number_format = NUM_FMT
    ws.cell(row=total_row, column=9, value=sum(li.get("labor_cost", 0) for li in line_items)).number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=10, value=sum(li.get("material_cost", 0) for li in line_items)).number_format = CURRENCY_FMT
    ws.cell(row=total_row, column=11, value=sum(li.get("subtotal", 0) for li in line_items)).number_format = CURRENCY_FMT
    style_total_row(ws, total_row, 12)

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["L"].width = 30


def export_estimate(estimate, output_path):
    """Export estimate to Excel workbook."""
    wb = openpyxl.Workbook()

    write_summary_sheet(wb, estimate)

    # Group line items by category/area
    categories = {}
    for li in estimate.get("line_items", []):
        cat = li.get("area", "Uncategorized")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(li)

    for cat_name, items in categories.items():
        safe_name = cat_name.replace("/", "-").replace("\\", "-").replace("*", "").replace("[", "").replace("]", "").replace(":", "").replace("?", "")[:31]
        write_detail_sheet(wb, safe_name, items)

    wb.save(output_path)
    return str(output_path)


def main():
    parser = argparse.ArgumentParser(description="CCF Estimate Excel Exporter")
    parser.add_argument("--estimate", required=True, help="Path to estimate JSON")
    parser.add_argument("--output", required=True, help="Output Excel file path")
    args = parser.parse_args()

    with open(args.estimate) as f:
        estimate = json.load(f)

    result = export_estimate(estimate, args.output)
    print(json.dumps({"exported": result}))


if __name__ == "__main__":
    main()
