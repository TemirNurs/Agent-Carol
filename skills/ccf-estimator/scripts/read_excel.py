#!/usr/bin/env python3
"""
CCF Pricing Workbook Reader
Reads CCF-Pricing-Policy-Production-Rates.xlsx and outputs structured JSON.
Usage: python read_excel.py --file <path> [--section all|rates|pricing|markup]
"""

import argparse
import json
import sys
import re
from pathlib import Path

import openpyxl


def parse_dollar_range(val):
    """Parse '$0.40' or '$0.55-0.75' into (low, high) floats."""
    if val is None:
        return None
    s = str(val).strip().replace("$", "")
    if "-" in s:
        parts = s.split("-")
        return (float(parts[0]), float(parts[1]))
    try:
        return (float(s), float(s))
    except ValueError:
        return None


def parse_pct_range(val):
    """Parse '15-20%' or '13-17%' into (low, high) floats as decimals."""
    if val is None:
        return None
    s = str(val).strip().replace("%", "")
    if "-" in s:
        parts = s.split("-")
        return (float(parts[0]) / 100, float(parts[1]) / 100)
    try:
        return (float(s) / 100, float(s) / 100)
    except ValueError:
        return None


def read_production_rates(wb):
    """Parse 'Production Rates' sheet into structured data."""
    ws = wb["Production Rates"]
    rates = {"painting": [], "prep": [], "non_painting": []}
    current_section = "painting"
    current_category = ""

    for row in ws.iter_rows(min_row=6, max_row=ws.max_row, values_only=False):
        cells = {}
        for c in row:
            try:
                cells[c.column_letter] = c.value
            except AttributeError:
                continue
        task = cells.get("B")
        if not task:
            continue

        task = str(task).strip()

        # Section headers
        if task.startswith("PREP WORK"):
            current_section = "prep"
            continue
        if task.startswith("NON-PAINTING"):
            current_section = "non_painting"
            continue

        # Category headers (all caps, no indent)
        if task == task.upper() and not task.startswith(" ") and cells.get("C") is None:
            current_category = task
            continue

        # Column header rows
        if task in ("Task / Surface", "Task / Trade"):
            continue

        unit = cells.get("C")
        if unit is None:
            continue

        unit = str(unit).strip()

        if current_section == "non_painting":
            low = cells.get("D")
            market = cells.get("E")
            notes = cells.get("F", "")
            rates["non_painting"].append({
                "task": task.strip(),
                "unit": unit,
                "low_rate": str(low) if low else None,
                "market_rate": str(market) if market else None,
                "notes": str(notes) if notes else "",
            })
        else:
            slow = cells.get("D")
            avg = cells.get("E")
            fast = cells.get("F")
            coats = cells.get("G")
            notes = cells.get("H", "")

            entry = {
                "task": task.strip(),
                "category": current_category,
                "unit": unit,
                "slow_rate": slow,
                "avg_rate": avg,
                "fast_rate": fast,
                "coats": str(coats) if coats else None,
                "notes": str(notes) if notes else "",
            }
            rates[current_section].append(entry)

    return rates


def read_pricing_policy(wb):
    """Parse 'Pricing Policy' sheet into structured pricing data."""
    ws = wb["Pricing Policy"]
    pricing = {"labor_rates": [], "unit_prices": [], "markup_scenarios": [], "overhead": [], "bid_formula": []}
    current_section = ""

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        cells = {}
        for c in row:
            try:
                cells[c.column_letter] = c.value
            except AttributeError:
                continue
        b_val = cells.get("B")
        if not b_val:
            continue
        b_val = str(b_val).strip()

        # Section detection
        if "LABOR RATE STRUCTURE" in b_val:
            current_section = "labor"
            continue
        if "COMPETITIVE $/SF PRICING" in b_val:
            current_section = "pricing"
            continue
        if "MARKUP & MARGIN POLICY" in b_val:
            current_section = "markup"
            continue
        if "OVERHEAD RATE" in b_val:
            current_section = "overhead"
            continue
        if "BID PRICING FORMULA" in b_val:
            current_section = "formula"
            continue

        # Skip headers/insights
        if b_val.startswith("KEY INSIGHT") or b_val.startswith("Role") or b_val.startswith("Scope / Task") or b_val.startswith("Scenario") or b_val.startswith("Overhead Item"):
            continue
        if b_val.startswith("Our overhead"):
            continue

        if current_section == "labor":
            c_val = cells.get("C")
            if c_val and "$" in str(c_val):
                pricing["labor_rates"].append({
                    "role": b_val,
                    "hourly_wage": str(c_val),
                    "burden_pct": str(cells.get("D", "")),
                    "burdened_rate": str(cells.get("E", "")),
                    "market_comparison": str(cells.get("F", "")),
                })

        elif current_section == "pricing":
            c_val = cells.get("C")
            if c_val and "$" in str(c_val):
                floor_price = parse_dollar_range(cells.get("D"))
                target_price = parse_dollar_range(cells.get("E"))
                premium_price = parse_dollar_range(cells.get("F"))
                pricing["unit_prices"].append({
                    "task": b_val.strip(),
                    "unit": str(c_val),
                    "floor_price": floor_price,
                    "target_price": target_price,
                    "premium_price": premium_price,
                    "when_to_use": str(cells.get("G", "")),
                })
            elif b_val in ("INTERIOR PAINTING", "EXTERIOR PAINTING", "SPECIALTY"):
                continue

        elif current_section == "markup":
            c_val = cells.get("C")
            if c_val and "%" in str(c_val):
                markup_range = parse_pct_range(c_val)
                margin_range = parse_pct_range(cells.get("D"))
                pricing["markup_scenarios"].append({
                    "scenario": b_val,
                    "markup_range": markup_range,
                    "margin_range": margin_range,
                    "when_to_apply": str(cells.get("E", "")),
                })

        elif current_section == "overhead":
            c_val = cells.get("C")
            if c_val and ("$" in str(c_val) or "In" in str(c_val)):
                pricing["overhead"].append({
                    "item": b_val,
                    "monthly": str(c_val),
                    "annual": str(cells.get("D", "")),
                    "pct_of_revenue": str(cells.get("E", "")),
                })

        elif current_section == "formula":
            if b_val.startswith("STEP") or b_val.startswith("EXAMPLE"):
                pricing["bid_formula"].append(b_val)

    return pricing


def read_all(filepath):
    """Read entire pricing workbook and return all data."""
    wb = openpyxl.load_workbook(filepath, data_only=False)
    return {
        "production_rates": read_production_rates(wb),
        "pricing_policy": read_pricing_policy(wb),
        "source_file": str(filepath),
    }


def read_takeoff_file(filepath):
    """Read a takeoff CSV or Excel file and return structured line items."""
    path = Path(filepath)
    import pandas as pd

    if path.suffix.lower() == ".csv":
        df = pd.read_csv(filepath)
    else:
        df = pd.read_excel(filepath)

    # Normalize column names
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    items = []
    for _, row in df.iterrows():
        item = {}
        for col in df.columns:
            val = row[col]
            if pd.notna(val):
                item[col] = val if not isinstance(val, float) or val != int(val) else int(val)
            else:
                item[col] = None
        items.append(item)

    return items


def main():
    parser = argparse.ArgumentParser(description="CCF Pricing Workbook Reader")
    parser.add_argument("--file", required=True, help="Path to Excel/CSV file")
    parser.add_argument("--section", default="all", choices=["all", "rates", "pricing", "takeoff"],
                        help="Which section to read")
    parser.add_argument("--output", default="json", choices=["json"], help="Output format")
    args = parser.parse_args()

    filepath = Path(args.file)
    if not filepath.exists():
        print(json.dumps({"error": f"File not found: {filepath}"}), file=sys.stderr)
        sys.exit(1)

    if args.section == "takeoff":
        data = read_takeoff_file(filepath)
    else:
        wb = openpyxl.load_workbook(filepath, data_only=False)
        if args.section == "rates":
            data = read_production_rates(wb)
        elif args.section == "pricing":
            data = read_pricing_policy(wb)
        else:
            data = read_all(filepath)

    print(json.dumps(data, indent=2, default=str))


if __name__ == "__main__":
    main()
