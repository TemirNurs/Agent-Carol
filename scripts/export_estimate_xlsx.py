#!/usr/bin/env python3
"""Export estimate to Excel (.xlsx) matching CCF template layout exactly.

Template source: Boot Barn, Food Lion, Venture IV estimates.
Layout: 6 sheets — Estimate Summary, Interior Walls, Ceilings, Doors & Frames, Exterior, Prep & Misc
Style: Georgia titles, Calibri data, navy #0D1B2A headers, alternating gray #F3F4F6 rows,
       green #008000 bid price on #E8F5E9 background.
"""
import json, sys, os
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).resolve().parent.parent
PROJECTS_DIR = BASE_DIR / "data" / "projects"

# ── CCF Template Colors (from Boot Barn / Food Lion / Venture IV templates) ──
NAVY = "0D1B2A"
ALT_ROW = "F3F4F6"
BID_GREEN = "008000"
BID_GREEN_BG = "E8F5E9"
DATA_TEXT = "1F2937"
NOTE_TEXT = "6B7280"

# ── Reusable styles ──
TITLE_FONT = Font(name="Georgia", size=14, bold=True, color=NAVY)
SECTION_FONT = Font(name="Georgia", size=11, bold=True, color=NAVY)
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
TOTALS_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TOTALS_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10, color=DATA_TEXT)
DATA_FONT_B = Font(name="Calibri", size=10, bold=True, color=DATA_TEXT)
NOTE_FONT = Font(name="Calibri", size=9, color=NOTE_TEXT)
INFO_FONT = Font(name="Calibri", size=9, color=NOTE_TEXT)
BID_FONT = Font(name="Georgia", size=12, bold=True, color=BID_GREEN)
BID_FILL = PatternFill(start_color=BID_GREEN_BG, end_color=BID_GREEN_BG, fill_type="solid")
GREEN_FONT = Font(name="Calibri", size=10, bold=True, color=BID_GREEN)
ALT_FILL = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
ALIGN_L = Alignment(horizontal="left", vertical="center")
ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_R = Alignment(horizontal="right", vertical="center")

MONEY_FMT = "\\$#,##0.00"
HRS_FMT = "#,##0.0"
PCT_FMT = "0.0%"
QTY_FMT = "#,##0"

# Illustrative default — real rates come from the gitignored pricing config.
LABOR_RATE = 28.00


def _header_row(ws, row, headers, col_start=1):
    """Write a navy header row."""
    for c, h in enumerate(headers, col_start):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_C


def _totals_row(ws, row, values, col_start=1):
    """Write a navy totals row with white text."""
    for c, v in enumerate(values, col_start):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = TOTALS_FONT
        cell.fill = TOTALS_FILL
        cell.alignment = ALIGN_L if c == col_start else ALIGN_R


def _data_cell(ws, row, col, value, font=None, fmt=None, align=None, alt=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or DATA_FONT
    if fmt:
        cell.number_format = fmt
    cell.alignment = align or ALIGN_L
    if alt:
        cell.fill = ALT_FILL
    return cell


def _alt_row_fill(ws, row, max_col, is_alt):
    if is_alt:
        for c in range(1, max_col + 1):
            ws.cell(row=row, column=c).fill = ALT_FILL


def _note_row(ws, row, text, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = NOTE_FONT
    cell.alignment = ALIGN_L


def _title_row(ws, row, text, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = TITLE_FONT
    cell.alignment = ALIGN_L
    ws.row_dimensions[row].height = 24


def _section_label(ws, row, text, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.alignment = ALIGN_L


# ── Classify items into categories ──
def _classify_items(items, takeoff_items):
    """Classify items into Interior Walls, Ceilings, Doors, Exterior, Wallcovering, Prep."""
    tk_lookup = {}
    for ti in takeoff_items:
        tk_lookup[ti["area"]] = ti

    interior = []
    ceilings = []
    doors = []
    exterior = []
    wallcovering = []
    prep = []

    for item in items:
        area = item["area"].lower()
        tk = tk_lookup.get(item["area"], {})
        task = tk.get("task_code", "")

        if "ceiling" in area or "roof structure" in area or "metal_deck" in task or "ceiling" in task:
            ceilings.append((item, tk))
        elif "wallcovering" in area or "wallcovering" in task:
            wallcovering.append((item, tk))
        elif "pressure" in area or "pressure" in task:
            prep.append((item, tk))
        elif any(w in area for w in ["canopy", "bollard", "trash", "vent", "fuel"]) or \
             any(w in task for w in ["exterior", "fuel_equipment"]) or \
             ("exterior" in area and "door" not in area):
            exterior.append((item, tk))
        elif "exterior" in area and "door" in area:
            # Exterior doors go in Exterior sheet, not Doors
            exterior.append((item, tk))
        elif "door" in area or "door" in task:
            doors.append((item, tk))
        else:
            interior.append((item, tk))

    return interior, ceilings, doors, exterior, wallcovering, prep


def export_estimate(slug: str) -> str:
    """Export estimate.json to a multi-sheet Excel workbook matching CCF template."""
    proj_dir = PROJECTS_DIR / slug
    estimate_path = proj_dir / "estimate.json"
    takeoff_path = proj_dir / "takeoff.json"

    if not estimate_path.exists():
        return f"ERROR: No estimate found at {estimate_path}"

    est = json.loads(estimate_path.read_text(encoding="utf-8"))
    takeoff = {}
    if takeoff_path.exists():
        takeoff = json.loads(takeoff_path.read_text(encoding="utf-8"))

    summary = est.get("summary", {})
    # Schema shim: producers (estimator_agent, build_estimate) emit
    # est["line_items"] with labor_cost/material_cost keys; this exporter's
    # sheets were written against est["items"] with labor/material. Without
    # this mapping every detail sheet exported EMPTY while the bid price
    # showed — a customer-facing workbook with no backup.
    items = est.get("items") or est.get("line_items") or []
    norm = []
    for it in items:
        if not isinstance(it, dict):
            continue
        d = dict(it)
        d.setdefault("labor", d.get("labor_cost", 0) or 0)
        d.setdefault("material", d.get("material_cost", 0) or 0)
        d.setdefault("desc", d.get("description") or d.get("name") or "")
        d.setdefault("unit", d.get("uom") or d.get("unit") or "")
        d.setdefault("qty", d.get("quantity", d.get("qty", 0)) or 0)
        # _classify_items routes sheets by item["area"]
        d.setdefault("area", d.get("category") or d.get("desc") or "")
        # line_total: producers emit ext/total/line_total inconsistently
        if "line_total" not in d:
            d["line_total"] = (d.get("ext") or d.get("total")
                               or (d["labor"] + d["material"]) or 0)
        norm.append(d)
    items = norm
    tk_items = takeoff.get("items", [])
    interior, ceilings, doors, exterior, wallcovering, prep = _classify_items(items, tk_items)

    project_name = est.get("project", slug.replace("_", " ").title())
    bid_date = datetime.now().strftime("%B %d, %Y")
    oh_pct = est.get("oh_pct", 0.12)
    profit_pct = est.get("profit_pct", 0.18)

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════
    # SHEET 1: Estimate Summary
    # ═══════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Estimate Summary"
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14

    r = 1
    # Row 1: Title
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="PAINTING & WALLCOVERING ESTIMATE")
    c.font = Font(name="Georgia", size=14, bold=True, color=NAVY)
    c.alignment = ALIGN_L
    ws.row_dimensions[r].height = 24
    r += 1

    # Row 2: Project name
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value=project_name)
    c.font = Font(name="Georgia", size=11, bold=True, color=NAVY)
    c.alignment = ALIGN_L
    r += 1

    # Row 3: Project location — from the estimate data, NEVER hardcoded
    # (a leftover "N. Myrtle Beach, SC" template literal was stamping the
    # wrong address onto every customer-facing workbook).
    loc = est.get("address") or ", ".join(
        x for x in (est.get("city"), est.get("state")) if x) or ""
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value=loc)
    c.font = INFO_FONT
    c.alignment = ALIGN_L
    r += 1

    # Row 4: Bid date / prepared by (company from config)
    from _lib import company as _co
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value=f"Bid Date: {bid_date}  |  Prepared by: {_co.company_name()}")
    c.font = INFO_FONT
    c.alignment = ALIGN_L
    r += 1

    # Row 5: Company info
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value="Budget Painting and Wallcovering LLC  |  3308 Chancellor Lane, Monroe, NC 28110  |  (980) 348-1827")
    c.font = INFO_FONT
    c.alignment = ALIGN_L
    r += 1

    # Row 6: Date prepared
    ws.merge_cells(f"A{r}:F{r}")
    c = ws.cell(row=r, column=1, value=f"Date Prepared: {bid_date} — DRAWING-DIMENSION ESTIMATE")
    c.font = INFO_FONT
    c.alignment = ALIGN_L
    r += 1

    r += 1  # blank row

    # Header row
    _header_row(ws, r, ["SCOPE CATEGORY", "LABOR HRS", "LABOR $", "MATERIALS $", "DIRECT COST", "% OF TOTAL"])
    r += 1

    # Build category summaries
    def _cat_totals(cat_items):
        labor = sum(i["labor"] for i, _ in cat_items)
        mat = sum(i["material"] for i, _ in cat_items)
        hrs = labor / LABOR_RATE if LABOR_RATE else 0
        return hrs, labor, mat, labor + mat

    cats = []
    if interior:
        h, l, m, dc = _cat_totals(interior)
        cats.append(("Interior Walls", h, l, m, dc))
    if ceilings:
        h, l, m, dc = _cat_totals(ceilings)
        cats.append(("Ceilings", h, l, m, dc))
    if doors:
        h, l, m, dc = _cat_totals(doors)
        cats.append((f"Doors & Frames ({sum(i['quantity'] for i, _ in doors)} units)", h, l, m, dc))
    if exterior:
        h, l, m, dc = _cat_totals(exterior)
        cats.append(("Exterior Painting", h, l, m, dc))
    if wallcovering:
        h, l, m, dc = _cat_totals(wallcovering)
        cats.append(("Wallcovering", h, l, m, dc))
    if prep:
        h, l, m, dc = _cat_totals(prep)
        cats.append(("Prep & Misc", h, l, m, dc))

    # Equipment & mobilization as a prep/misc line if not already in prep
    equip_mob = summary.get("equipment", 0) + summary.get("mobilization", 0)
    if equip_mob > 0 and not prep:
        cats.append(("Equipment & Mobilization", 0, 0, equip_mob, equip_mob))

    total_dc = sum(c[4] for c in cats)
    data_start = r
    for idx, (name, hrs, labor, mat, dc) in enumerate(cats):
        is_alt = idx % 2 == 1
        _data_cell(ws, r, 1, name, DATA_FONT, align=ALIGN_L, alt=is_alt)
        _data_cell(ws, r, 2, hrs, DATA_FONT, HRS_FMT, ALIGN_R, is_alt)
        _data_cell(ws, r, 3, labor, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws, r, 4, mat, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws, r, 5, dc, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        pct = dc / total_dc if total_dc else 0
        _data_cell(ws, r, 6, pct, DATA_FONT, PCT_FMT, ALIGN_R, is_alt)
        r += 1

    # Direct cost totals row (navy)
    total_hrs = summary.get("labor_hours", sum(c[1] for c in cats))
    total_labor = summary.get("direct_labor", sum(c[2] for c in cats))
    total_mat = summary.get("materials", sum(c[3] for c in cats))
    total_direct = summary.get("direct_cost", total_dc)
    vals = ["DIRECT COST TOTALS", total_hrs, total_labor, total_mat, total_direct, 1]
    _totals_row(ws, r, vals)
    ws.cell(row=r, column=2).number_format = HRS_FMT
    ws.cell(row=r, column=3).number_format = MONEY_FMT
    ws.cell(row=r, column=4).number_format = MONEY_FMT
    ws.cell(row=r, column=5).number_format = MONEY_FMT
    ws.cell(row=r, column=6).number_format = PCT_FMT
    r += 2

    # Bid calculation
    overhead = summary.get("overhead", total_direct * oh_pct)
    markup = summary.get("profit", (total_direct + overhead) * profit_pct)
    equip = summary.get("equipment", 0)
    mob = summary.get("mobilization", 0)
    bid_price = summary.get("total_bid", total_direct + overhead + markup + equip + mob)

    for label, val in [("Direct Cost", total_direct),
                       (f"Overhead ({oh_pct*100:.0f}%)", overhead),
                       (f"Markup ({profit_pct*100:.0f}%)", markup)]:
        _data_cell(ws, r, 1, label, DATA_FONT_B, align=ALIGN_L)
        _data_cell(ws, r, 5, val, GREEN_FONT, MONEY_FMT, ALIGN_R)
        r += 1
    if equip > 0:
        _data_cell(ws, r, 1, "Equipment (Lift Rental)", DATA_FONT_B, align=ALIGN_L)
        _data_cell(ws, r, 5, equip, GREEN_FONT, MONEY_FMT, ALIGN_R)
        r += 1
    if mob > 0:
        _data_cell(ws, r, 1, "Mobilization", DATA_FONT_B, align=ALIGN_L)
        _data_cell(ws, r, 5, mob, GREEN_FONT, MONEY_FMT, ALIGN_R)
        r += 1

    # BID PRICE row (green on light green bg)
    c = ws.cell(row=r, column=1, value="BID PRICE (PAINTING & WALLCOVERING)")
    c.font = BID_FONT
    c.fill = BID_FILL
    c.alignment = ALIGN_L
    for col in range(2, 7):
        ws.cell(row=r, column=col).fill = BID_FILL
    c = ws.cell(row=r, column=5, value=bid_price)
    c.font = BID_FONT
    c.fill = BID_FILL
    c.number_format = MONEY_FMT
    c.alignment = ALIGN_R
    ws.row_dimensions[r].height = 21.75
    r += 2

    # KEY METRICS section
    _section_label(ws, r, "KEY METRICS", 6)
    r += 1
    total_sf = sum(i["quantity"] for i in items if i["unit"] == "SF")
    metrics = [
        ("Total Paintable SF:", total_sf, QTY_FMT, DATA_FONT, GREEN_FONT),
        ("Blended Rate ($/SF):", bid_price / total_sf if total_sf else 0, MONEY_FMT, DATA_FONT, GREEN_FONT),
        ("Total Labor Hours:", total_hrs, QTY_FMT, DATA_FONT, GREEN_FONT),
        ("Estimated Duration (3-man crew):", f"{summary.get('duration_days', 0)} working days", None, DATA_FONT, DATA_FONT_B),
        ("Labor Rate (burdened):", f"${LABOR_RATE:.2f}/hr", None, DATA_FONT, DATA_FONT_B),
    ]
    for label, val, fmt, lfont, vfont in metrics:
        _data_cell(ws, r, 1, label, lfont, align=ALIGN_L)
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = vfont
        if fmt:
            vc.number_format = fmt
        r += 1

    r += 1

    # EXCLUSIONS
    excls = est.get("exclusions", [])
    if excls:
        _section_label(ws, r, "EXCLUSIONS", 6)
        r += 1
        for ex in excls:
            _note_row(ws, r, f"  \u2022  {ex}", 6)
            r += 1

    r += 1

    # NOTES
    _section_label(ws, r, "NOTES", 6)
    r += 1
    notes = [
        "All paint Sherwin-Williams products (National Account pricing)",
        "Daytime work assumed unless otherwise directed by GC",
        "One mobilization included; additional mob/demob at cost",
        "All new construction — paint systems per A6.1 Materials Schedule",
        f"Estimate based on drawing dimensions ({est.get('methodology', '')})",
    ]
    for n in notes:
        _note_row(ws, r, f"  \u2022  {n}", 6)
        r += 1

    # ═══════════════════════════════════════════════════════════════
    # SHEET 2: Interior Walls
    # ═══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Interior Walls")
    mc = 12  # max col (A thru L)
    ws2.column_dimensions["A"].width = 8
    ws2.column_dimensions["B"].width = 36
    ws2.column_dimensions["C"].width = 12
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 12
    ws2.column_dimensions["F"].width = 10
    ws2.column_dimensions["G"].width = 8
    ws2.column_dimensions["H"].width = 42
    ws2.column_dimensions["I"].width = 12
    ws2.column_dimensions["J"].width = 14
    ws2.column_dimensions["K"].width = 12
    ws2.column_dimensions["L"].width = 14

    r = 1
    _title_row(ws2, r, "INTERIOR WALLS — BREAKDOWN", mc)
    r += 2

    _header_row(ws2, r, ["#", "Room / Area", "Takeoff SF", "SOW %", "Paint SF", "Method",
                          "Coats", "Paint System", "Labor Hrs", "Labor $", "Material $", "Total $"])
    r += 1

    tot_tsf = tot_psf = tot_hrs = tot_lab = tot_mat = tot_total = 0
    for idx, (item, tk) in enumerate(interior):
        is_alt = idx % 2 == 1
        qty = item["quantity"]
        labor = item["labor"]
        mat = item["material"]
        total = item["line_total"]
        hrs = labor / LABOR_RATE
        method = tk.get("method", item.get("method", "")).replace("_", " ").upper()
        if method == "BRUSH ROLL":
            method = "B&R"
        coats = tk.get("coats", "P+2")
        color = item.get("color", "")
        product = tk.get("product", "")
        paint_sys = f"{product} — {color}" if product else color

        _data_cell(ws2, r, 1, idx + 1, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws2, r, 2, item["area"], DATA_FONT, align=ALIGN_L, alt=is_alt)
        _data_cell(ws2, r, 3, qty, DATA_FONT, QTY_FMT, ALIGN_R, is_alt)
        _data_cell(ws2, r, 4, "100%", DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws2, r, 5, qty, DATA_FONT, QTY_FMT, ALIGN_R, is_alt)
        _data_cell(ws2, r, 6, method, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws2, r, 7, coats, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws2, r, 8, paint_sys, DATA_FONT, align=ALIGN_L, alt=is_alt)
        _data_cell(ws2, r, 9, hrs, DATA_FONT, HRS_FMT, ALIGN_R, is_alt)
        _data_cell(ws2, r, 10, labor, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws2, r, 11, mat, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws2, r, 12, total, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)

        tot_tsf += qty
        tot_psf += qty
        tot_hrs += hrs
        tot_lab += labor
        tot_mat += mat
        tot_total += total
        r += 1

    # Totals
    vals = ["", "INTERIOR WALLS — TOTAL", tot_tsf, "", tot_psf, "", "", "",
            tot_hrs, tot_lab, tot_mat, tot_total]
    _totals_row(ws2, r, vals)
    ws2.cell(row=r, column=3).number_format = QTY_FMT
    ws2.cell(row=r, column=5).number_format = QTY_FMT
    ws2.cell(row=r, column=9).number_format = HRS_FMT
    ws2.cell(row=r, column=10).number_format = MONEY_FMT
    ws2.cell(row=r, column=11).number_format = MONEY_FMT
    ws2.cell(row=r, column=12).number_format = MONEY_FMT

    # ═══════════════════════════════════════════════════════════════
    # SHEET 3: Ceilings
    # ═══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Ceilings")
    mc3 = 10
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 14
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 10
    ws3.column_dimensions["F"].width = 42
    ws3.column_dimensions["G"].width = 12
    ws3.column_dimensions["H"].width = 14
    ws3.column_dimensions["I"].width = 12
    ws3.column_dimensions["J"].width = 14

    r = 1
    _title_row(ws3, r, "CEILINGS — BY TYPE & AREA", mc3)
    r += 1
    _header_row(ws3, r, ["Area / Room", "Ceiling Type", "Total SF", "Prod Rate (SF/hr)",
                          "Coats", "Paint System", "Labor Hrs", "Labor $", "Material $", "Total $"])
    r += 1

    tot_sf3 = tot_hrs3 = tot_lab3 = tot_mat3 = tot_total3 = 0
    for idx, (item, tk) in enumerate(ceilings):
        is_alt = idx % 2 == 1
        qty = item["quantity"]
        labor = item["labor"]
        mat = item["material"]
        total = item["line_total"]
        hrs = labor / LABOR_RATE
        ctype = "GWB Ceiling" if "gwb" in item["area"].lower() or "ceiling" in item["area"].lower() else "Exposed Structure/Deck"
        if "roof" in item["area"].lower() or "metal" in item["area"].lower():
            ctype = "Exposed Structure/Deck (Dryfall)"
        product = tk.get("product", "")
        color = item.get("color", "")
        paint_sys = f"{product} — {color}" if product else color
        coats = tk.get("coats", "P+2")
        prod_rate = 110 if "exposed" in ctype.lower() or "dryfall" in ctype.lower() else 200

        _data_cell(ws3, r, 1, item["area"], DATA_FONT, align=ALIGN_L, alt=is_alt)
        _data_cell(ws3, r, 2, ctype, DATA_FONT, align=ALIGN_L, alt=is_alt)
        _data_cell(ws3, r, 3, qty, DATA_FONT, QTY_FMT, ALIGN_R, is_alt)
        _data_cell(ws3, r, 4, prod_rate, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws3, r, 5, coats, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws3, r, 6, paint_sys, DATA_FONT, align=ALIGN_L, alt=is_alt)
        _data_cell(ws3, r, 7, hrs, DATA_FONT, HRS_FMT, ALIGN_R, is_alt)
        _data_cell(ws3, r, 8, labor, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws3, r, 9, mat, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws3, r, 10, total, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)

        tot_sf3 += qty
        tot_hrs3 += hrs
        tot_lab3 += labor
        tot_mat3 += mat
        tot_total3 += total
        r += 1

    vals = ["CEILING TOTALS", "", tot_sf3, "", "", "", tot_hrs3, tot_lab3, tot_mat3, tot_total3]
    _totals_row(ws3, r, vals)
    ws3.cell(row=r, column=3).number_format = QTY_FMT
    ws3.cell(row=r, column=7).number_format = HRS_FMT
    ws3.cell(row=r, column=8).number_format = MONEY_FMT
    ws3.cell(row=r, column=9).number_format = MONEY_FMT
    ws3.cell(row=r, column=10).number_format = MONEY_FMT
    r += 2

    _section_label(ws3, r, "NOTES", mc3)
    r += 1
    for n in [
        "GWB ceiling soffits/bulkheads in sales area per A2.6 Reflected Ceiling Plan",
        "Exposed structure rate 110 SF/hr accounts for irregular surfaces, spray from lift",
        "Scissor/boom lift required — rental cost on Prep & Misc section",
    ]:
        _note_row(ws3, r, f"  \u2022  {n}", mc3)
        r += 1

    # ═══════════════════════════════════════════════════════════════
    # SHEET 4: Doors & Frames
    # ═══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Doors & Frames")
    mc4 = 9
    ws4.column_dimensions["A"].width = 8
    ws4.column_dimensions["B"].width = 36
    ws4.column_dimensions["C"].width = 8
    ws4.column_dimensions["D"].width = 10
    ws4.column_dimensions["E"].width = 11
    ws4.column_dimensions["F"].width = 13
    ws4.column_dimensions["G"].width = 11
    ws4.column_dimensions["H"].width = 13
    ws4.column_dimensions["I"].width = 13

    r = 1
    _title_row(ws4, r, "DOORS & FRAMES — ITEMIZED", mc4)
    r += 1
    _header_row(ws4, r, ["Item #", "Description", "Qty", "Hrs/Unit", "Total Hrs",
                          "Labor $", "Mat $/Unit", "Total Mat $", "Total $"])
    r += 1

    tot_qty4 = tot_hrs4 = tot_lab4 = tot_mat4 = tot_total4 = 0
    for idx, (item, tk) in enumerate(doors):
        is_alt = idx % 2 == 1
        qty = item["quantity"]
        labor = item["labor"]
        mat = item["material"]
        total = item["line_total"]
        hrs = labor / LABOR_RATE
        hrs_per = hrs / qty if qty else 0
        mat_per = mat / qty if qty else 0

        _data_cell(ws4, r, 1, idx + 1, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws4, r, 2, item["area"], DATA_FONT, align=ALIGN_L, alt=is_alt)
        _data_cell(ws4, r, 3, qty, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws4, r, 4, hrs_per, DATA_FONT, HRS_FMT, ALIGN_R, is_alt)
        _data_cell(ws4, r, 5, hrs, DATA_FONT, HRS_FMT, ALIGN_R, is_alt)
        _data_cell(ws4, r, 6, labor, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws4, r, 7, mat_per, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws4, r, 8, mat, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws4, r, 9, total, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)

        tot_qty4 += qty
        tot_hrs4 += hrs
        tot_lab4 += labor
        tot_mat4 += mat
        tot_total4 += total
        r += 1

    vals = ["", "DOOR TOTALS", tot_qty4, "", tot_hrs4, tot_lab4, "", tot_mat4, tot_total4]
    _totals_row(ws4, r, vals)
    ws4.cell(row=r, column=5).number_format = HRS_FMT
    ws4.cell(row=r, column=6).number_format = MONEY_FMT
    ws4.cell(row=r, column=8).number_format = MONEY_FMT
    ws4.cell(row=r, column=9).number_format = MONEY_FMT
    r += 2

    _section_label(ws4, r, "NOTES", mc4)
    r += 1
    for n in [
        "All HM doors per A6.0 Door Schedule — Pro Industrial Pre-Cat DTM Semi-Gloss, 2 coats",
        "Includes door frames — hardware masked, not removed",
    ]:
        _note_row(ws4, r, f"  \u2022  {n}", mc4)
        r += 1

    # ═══════════════════════════════════════════════════════════════
    # SHEET 5: Exterior
    # ═══════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Exterior")
    mc5 = 9
    ws5.column_dimensions["A"].width = 8
    ws5.column_dimensions["B"].width = 40
    ws5.column_dimensions["C"].width = 12
    ws5.column_dimensions["D"].width = 10
    ws5.column_dimensions["E"].width = 12
    ws5.column_dimensions["F"].width = 12
    ws5.column_dimensions["G"].width = 14
    ws5.column_dimensions["H"].width = 12
    ws5.column_dimensions["I"].width = 14

    r = 1
    _title_row(ws5, r, "EXTERIOR PAINTING", mc5)
    r += 1
    _header_row(ws5, r, ["Item #", "Description", "Qty / SF", "Unit", "Prod Rate",
                          "Labor Hrs", "Labor $", "Material $", "Total $"])
    r += 1

    tot_sf5 = tot_hrs5 = tot_lab5 = tot_mat5 = tot_total5 = 0
    for idx, (item, tk) in enumerate(exterior):
        is_alt = idx % 2 == 1
        qty = item["quantity"]
        labor = item["labor"]
        mat = item["material"]
        total = item["line_total"]
        hrs = labor / LABOR_RATE
        unit = item["unit"]
        method = tk.get("method", item.get("method", "")).replace("_", " ")

        _data_cell(ws5, r, 1, idx + 1, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws5, r, 2, item["area"], DATA_FONT, align=ALIGN_L, alt=is_alt)
        _data_cell(ws5, r, 3, qty, DATA_FONT, QTY_FMT if unit == "SF" else None, ALIGN_R, is_alt)
        _data_cell(ws5, r, 4, unit, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws5, r, 5, 150 if "spray" in method else 40, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws5, r, 6, hrs, DATA_FONT, HRS_FMT, ALIGN_R, is_alt)
        _data_cell(ws5, r, 7, labor, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws5, r, 8, mat, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws5, r, 9, total, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)

        tot_sf5 += qty if unit == "SF" else 0
        tot_hrs5 += hrs
        tot_lab5 += labor
        tot_mat5 += mat
        tot_total5 += total
        r += 1

    vals = ["", "EXTERIOR TOTALS", tot_sf5, "", "", tot_hrs5, tot_lab5, tot_mat5, tot_total5]
    _totals_row(ws5, r, vals)
    ws5.cell(row=r, column=3).number_format = QTY_FMT
    ws5.cell(row=r, column=6).number_format = HRS_FMT
    ws5.cell(row=r, column=7).number_format = MONEY_FMT
    ws5.cell(row=r, column=8).number_format = MONEY_FMT
    ws5.cell(row=r, column=9).number_format = MONEY_FMT
    r += 2

    _section_label(ws5, r, "NOTES", mc5)
    r += 1
    for n in [
        "All exterior DTM: Sher-Cryl HPA Semi-Gloss DTM per A6.1 paint schedule",
        "Gas canopy requires boom lift — rental in Prep & Misc",
        "Fuel fill lids color-coded per fuel type (P-23 thru P-34)",
        "Spray application with backroll where applicable",
    ]:
        _note_row(ws5, r, f"  \u2022  {n}", mc5)
        r += 1

    # ═══════════════════════════════════════════════════════════════
    # SHEET 6: Prep & Misc
    # ═══════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Prep & Misc")
    mc6 = 9
    ws6.column_dimensions["A"].width = 8
    ws6.column_dimensions["B"].width = 44
    ws6.column_dimensions["C"].width = 8
    ws6.column_dimensions["D"].width = 8
    ws6.column_dimensions["E"].width = 10
    ws6.column_dimensions["F"].width = 12
    ws6.column_dimensions["G"].width = 14
    ws6.column_dimensions["H"].width = 13
    ws6.column_dimensions["I"].width = 13

    r = 1
    _title_row(ws6, r, "PREP, MISC & EQUIPMENT", mc6)
    r += 1
    _header_row(ws6, r, ["Item #", "Description", "Qty", "Unit", "Rate/Hr",
                          "Labor Hrs", "Labor $", "Material $", "Total $"])
    r += 1

    # Build prep items from estimate data + fixed items
    prep_line_items = []

    # Pressure wash if present
    for item, tk in prep:
        hrs = item["labor"] / LABOR_RATE
        prep_line_items.append({
            "desc": item["area"],
            "qty": item["quantity"],
            "unit": item["unit"],
            "rate": LABOR_RATE,
            "hrs": hrs,
            "labor": item["labor"],
            "mat": item["material"],
            "total": item["line_total"],
        })

    # Standard prep items
    equip_val = summary.get("equipment", 0)
    mob_val = summary.get("mobilization", 0)

    standard_prep = [
        ("Surface Prep — Skim/Patch New GWB", 1, "LS", 4, 4 * LABOR_RATE, 0),
        ("Caulking — Joints & Transitions", 1, "LS", 3, 3 * LABOR_RATE, 0),
        ("Floor/Fixture Masking & Protection", 1, "LS", 4, 4 * LABOR_RATE, 0),
        ("Daily Setup/Cleanup", 1, "LS", 4, 4 * LABOR_RATE, 0),
        ("Touch-up / Punch List", 1, "LS", 4, 4 * LABOR_RATE, 0),
        ("Boom Lift Rental", 1, "LS", 0, 0, equip_val),
        ("Mobilization / Demobilization", 1, "LS", 0, 0, mob_val) if mob_val else None,
        ("Supplies (tape, plastic, drops, rollers, tips)", 1, "LS", 0, 0, 150),
    ]

    for sp in standard_prep:
        if sp is None:
            continue
        desc, qty, unit, hrs, labor, mat = sp
        prep_line_items.append({
            "desc": desc, "qty": qty, "unit": unit,
            "rate": LABOR_RATE if hrs > 0 else None,
            "hrs": hrs, "labor": labor, "mat": mat,
            "total": labor + mat,
        })

    tot_hrs6 = tot_lab6 = tot_mat6 = tot_total6 = 0
    for idx, pi in enumerate(prep_line_items):
        is_alt = idx % 2 == 1
        _data_cell(ws6, r, 1, idx + 1, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws6, r, 2, pi["desc"], DATA_FONT, align=ALIGN_L, alt=is_alt)
        _data_cell(ws6, r, 3, pi["qty"], DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws6, r, 4, pi["unit"], DATA_FONT, align=ALIGN_C, alt=is_alt)
        rate_str = f"${LABOR_RATE:.2f}" if pi.get("rate") else "\u2014"
        _data_cell(ws6, r, 5, rate_str, DATA_FONT, align=ALIGN_C, alt=is_alt)
        _data_cell(ws6, r, 6, pi["hrs"] if pi["hrs"] else 0, DATA_FONT, HRS_FMT, ALIGN_R, is_alt)
        _data_cell(ws6, r, 7, pi["labor"], DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws6, r, 8, pi["mat"], DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
        _data_cell(ws6, r, 9, pi["total"], DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)

        tot_hrs6 += pi["hrs"]
        tot_lab6 += pi["labor"]
        tot_mat6 += pi["mat"]
        tot_total6 += pi["total"]
        r += 1

    vals = ["", "PREP & MISC TOTALS", "", "", "", tot_hrs6, tot_lab6, tot_mat6, tot_total6]
    _totals_row(ws6, r, vals)
    ws6.cell(row=r, column=6).number_format = HRS_FMT
    ws6.cell(row=r, column=7).number_format = MONEY_FMT
    ws6.cell(row=r, column=8).number_format = MONEY_FMT
    ws6.cell(row=r, column=9).number_format = MONEY_FMT

    # ═══════════════════════════════════════════════════════════════
    # OPTIONAL: Wallcovering Sheet (if wallcovering items exist)
    # ═══════════════════════════════════════════════════════════════
    if wallcovering:
        ws7 = wb.create_sheet("Wallcovering")
        # Move after Doors & Frames
        wb.move_sheet("Wallcovering", offset=-1)
        mc7 = 9
        ws7.column_dimensions["A"].width = 8
        ws7.column_dimensions["B"].width = 44
        ws7.column_dimensions["C"].width = 12
        ws7.column_dimensions["D"].width = 10
        ws7.column_dimensions["E"].width = 14
        ws7.column_dimensions["F"].width = 11
        ws7.column_dimensions["G"].width = 13
        ws7.column_dimensions["H"].width = 13
        ws7.column_dimensions["I"].width = 13

        r = 1
        _title_row(ws7, r, "WALLCOVERING", mc7)
        r += 1
        _header_row(ws7, r, ["Item #", "Description", "SF / Qty", "Unit", "Prod Rate",
                              "Labor Hrs", "Labor $", "Material $", "Total $"])
        r += 1

        tot_hrs7 = tot_lab7 = tot_mat7 = tot_total7 = 0
        for idx, (item, tk) in enumerate(wallcovering):
            is_alt = idx % 2 == 1
            qty = item["quantity"]
            labor = item["labor"]
            mat = item["material"]
            total = item["line_total"]
            hrs = labor / LABOR_RATE

            _data_cell(ws7, r, 1, idx + 1, DATA_FONT, align=ALIGN_C, alt=is_alt)
            desc = f"{item['area']} — {item.get('color', '')}"
            _data_cell(ws7, r, 2, desc, DATA_FONT, align=ALIGN_L, alt=is_alt)
            _data_cell(ws7, r, 3, qty, DATA_FONT, QTY_FMT, ALIGN_R, is_alt)
            _data_cell(ws7, r, 4, item["unit"], DATA_FONT, align=ALIGN_C, alt=is_alt)
            _data_cell(ws7, r, 5, "40 SF/hr", DATA_FONT, align=ALIGN_C, alt=is_alt)
            _data_cell(ws7, r, 6, hrs, DATA_FONT, HRS_FMT, ALIGN_R, is_alt)
            _data_cell(ws7, r, 7, labor, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
            _data_cell(ws7, r, 8, mat, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)
            _data_cell(ws7, r, 9, total, DATA_FONT, MONEY_FMT, ALIGN_R, is_alt)

            tot_hrs7 += hrs
            tot_lab7 += labor
            tot_mat7 += mat
            tot_total7 += total
            r += 1

        vals = ["", "WALLCOVERING TOTALS", "", "", "", tot_hrs7, tot_lab7, tot_mat7, tot_total7]
        _totals_row(ws7, r, vals)
        ws7.cell(row=r, column=6).number_format = HRS_FMT
        ws7.cell(row=r, column=7).number_format = MONEY_FMT
        ws7.cell(row=r, column=8).number_format = MONEY_FMT
        ws7.cell(row=r, column=9).number_format = MONEY_FMT
        r += 2

        _section_label(ws7, r, "WALLCOVERING NOTES", mc7)
        r += 1
        for n in [
            "Type II vinyl wallcovering — 54\" commercial grade",
            "Production rate 40 SF/hr includes: layout, cutting, pasting, hanging, trimming, rolling seams",
            "Material pricing at $2.50/SF illustrative — real rates come from the gitignored pricing config; verify with manufacturer rep for exact roll pricing and waste factor",
        ]:
            _note_row(ws7, r, f"  \u2022  {n}", mc7)
            r += 1

    # ── Print settings ──
    for ws_sheet in wb.worksheets:
        ws_sheet.sheet_properties.pageSetUpPr = None
        ws_sheet.page_setup.orientation = "landscape"
        ws_sheet.page_setup.fitToWidth = 1
        ws_sheet.page_setup.fitToHeight = 0

    # Save
    xlsx_name = f"{slug}_estimate.xlsx"
    xlsx_path = proj_dir / xlsx_name
    wb.save(str(xlsx_path))
    return str(xlsx_path)


if __name__ == "__main__":
    slug = sys.argv[1] if len(sys.argv) > 1 else "7_eleven_42834_n_myrtle_beach_sc"
    path = export_estimate(slug)
    if path.startswith("ERROR"):
        print(path)
        sys.exit(1)
    print(f"Excel estimate saved: {path}")
    print(f"Size: {os.path.getsize(path) / 1024:.1f} KB")
